from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
import pandas as pd
import json
from core.utils import find_table_in_excel, concatenate_dataframes
from core import logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = FastAPI()

with open('config.json', 'r') as f:
    config = json.load(f)

sheet_name = config['sheet_name']
template_path = config['template_path']    
    

def format_and_save_openpyxl(result_df, sheet_name):
    """ Preserve the formatting of the original excel file and save the consolidated data
    
    Args:
        result_df (pd.DataFrame): Dataframe to be saved
        sheet_name (str): Name of the sheet in the excel file

    Raises:
        ValueError: If the sheet name does not exist in the workbook
        
    Returns:
        None
    """
    try:
        df = pd.read_excel('input/Consolidation_Assignment.xlsx', sheet_name=sheet_name, engine='openpyxl')
        _, row_num = find_table_in_excel(df)
        
        input_location = template_path
        output_location ='output/Consolidated_data.xlsx'
        workbook = load_workbook(input_location)
        
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
        
        start_row = row_num + 3
        
        for row_index, row in enumerate(dataframe_to_rows(result_df, index=False, header=False), start=start_row):
            for col_index, value in enumerate(row, start=1):  # openpyxl uses 1-indexing for columns
                sheet.cell(row=row_index, column=col_index, value=value)
        
        workbook.save(output_location)
        logger.log_message(message=f"EXEC: Consolidated file saved at {output_location}", level=0)
    except Exception as e:
        logger.log_message(message=f"ERROR: Saving consolidated file", level=1)
        raise e

@app.post("/files/")
async def consolidate_files(files: list[UploadFile] = File(...)):
    """ API to upload multiple excel files and consolidate them into a single file
    
    Args:
        files (list[UploadFile], optional): List of files to be uploaded. Defaults to File(...).
    Returns:
        FileResponse: Returns the consolidated file as a response
    """
    frames = []
    
    for file in files:
        contents = await file.read()
        
        if file.filename.endswith('.xlsx'):
            logger.log_message(message=f"EXEC: Reading file {file.filename}", level=0)
            df = pd.read_excel(contents, sheet_name=sheet_name, engine='openpyxl')
            df, _ = find_table_in_excel(df)
        # print(f"____________{df.shape}_____________")
            frames.append(df)
        else:
            logger.log_message(message=f"ERROR: File {file.filename} is not an excel file", level=1)
    
    result_df = concatenate_dataframes(frames, template_path)
    # print(f"_____________{result_df.shape}_____________")

    format_and_save_openpyxl(result_df, sheet_name)
    
    return FileResponse(
        'output/Consolidation_Assignment.xlsx',
        filename='output.xlsx',
        content_disposition_type='attachment',
        media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
